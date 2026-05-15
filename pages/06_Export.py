import streamlit as st
import pandas as pd
from pathlib import Path
import json
from io import BytesIO

from src.ui_utils import load_css, empty_state, PROCESSED_DIR

from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.drawing.image import Image as XLImage

st.set_page_config(page_title="Export", page_icon="📤", layout="wide", initial_sidebar_state="expanded")
load_css()

st.title("Export")
st.caption("Generate laporan_sentimen.xlsx dari hasil training terakhir.")

# ===== PATHS =====
nb_dir = PROCESSED_DIR / "naive_bayes"
tfidf_dir = PROCESSED_DIR / "tfidf"

metrics_path = nb_dir / "metrics.json"
report_csv = nb_dir / "classification_report.csv"
cm_img = nb_dir / "confusion_matrix.png"

preproc_path = PROCESSED_DIR / "dataset_preprocessing_non_empty.csv"
tfidf_info = tfidf_dir / "tfidf_info.json"

out_xlsx = PROCESSED_DIR / "laporan_sentimen.xlsx"


def clean_label_series(series: pd.Series) -> pd.Series:
    lab = series.astype(str).str.strip().str.lower()
    lab = lab.replace({"nan": "", "none": "", "null": ""})
    # ambil hanya label valid
    return lab[lab.isin(["positif", "netral", "negatif"])]

def add_df_to_sheet(ws, df: pd.DataFrame, start_row: int = 1, start_col: int = 1):
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start_row):
        for c_idx, value in enumerate(row, start_col):
            ws.cell(row=r_idx, column=c_idx, value=value)

def generate_report_bytes() -> bytes:
    wb = Workbook()

    # ===== Sheet 1: ringkasan =====
    ws_sum = wb.active
    ws_sum.title = "ringkasan"

    metrics = {}
    if metrics_path.exists():
        metrics = json.loads(metrics_path.read_text(encoding="utf-8"))

    tfidf_meta = {}
    if tfidf_info.exists():
        try:
            tfidf_meta = json.loads(tfidf_info.read_text(encoding="utf-8"))
        except Exception:
            tfidf_meta = {}

    rows = [
        ["Akurasi", metrics.get("accuracy", "")],
        ["Macro F1", metrics.get("macro_f1", "")],
        ["Weighted F1", metrics.get("weighted_f1", "")],
        ["N data uji", metrics.get("n_test", "")],
        ["Labels order", ", ".join(metrics.get("labels_order", [])) if isinstance(metrics.get("labels_order", []), list) else metrics.get("labels_order", "")],
        ["TF-IDF ngram_range", str(tfidf_meta.get("ngram_range", ""))],
        ["TF-IDF min_df", tfidf_meta.get("min_df", "")],
        ["TF-IDF max_df", tfidf_meta.get("max_df", "")],
        ["TF-IDF max_features", tfidf_meta.get("max_features", "")],
        ["Test size", tfidf_meta.get("test_size", "")],
        ["Stratify used", tfidf_meta.get("stratify_used", "")],
        ["Label counts (train split input)", str(tfidf_meta.get("label_counts", ""))],
    ]
    df_sum = pd.DataFrame(rows, columns=["Item", "Nilai"])
    add_df_to_sheet(ws_sum, df_sum)

    # ===== Sheet 2: distribusi label =====
    ws_dist = wb.create_sheet("distribusi_label")
    if preproc_path.exists():
        df_pre = pd.read_csv(preproc_path, encoding="utf-8-sig")
        if "label_manual" in df_pre.columns:
            lab = clean_label_series(df_pre["label_manual"])
            dist = lab.value_counts().reindex(["negatif", "netral", "positif"]).fillna(0).astype(int).reset_index()
            dist.columns = ["label", "jumlah"]
            add_df_to_sheet(ws_dist, dist)
        else:
            add_df_to_sheet(ws_dist, pd.DataFrame([["kolom label_manual tidak ada"]], columns=["info"]))
    else:
        add_df_to_sheet(ws_dist, pd.DataFrame([["dataset_preprocessing_non_empty.csv tidak ditemukan"]], columns=["info"]))

    # ===== Sheet 3: classification report =====
    ws_rep = wb.create_sheet("classification_report")
    if report_csv.exists():
        rep = pd.read_csv(report_csv, encoding="utf-8-sig")
        # rapihin kalau ada Unnamed: 0
        if "Unnamed: 0" in rep.columns:
            rep = rep.rename(columns={"Unnamed: 0": "label"})
        add_df_to_sheet(ws_rep, rep)
    else:
        add_df_to_sheet(ws_rep, pd.DataFrame([["classification_report.csv tidak ditemukan"]], columns=["info"]))

    # ===== Sheet 4: sample data =====
    ws_sample = wb.create_sheet("sample_data")
    if preproc_path.exists():
        df_pre = pd.read_csv(preproc_path, encoding="utf-8-sig")
        cols = [c for c in ["id_respon", "layanan", "komentar_teks", "teks_bersih", "label_manual"] if c in df_pre.columns]
        sample = df_pre[cols].head(200)
        add_df_to_sheet(ws_sample, sample)
    else:
        add_df_to_sheet(ws_sample, pd.DataFrame([["dataset_preprocessing_non_empty.csv tidak ditemukan"]], columns=["info"]))

    # ===== Sheet 5: confusion matrix image =====
    ws_cm = wb.create_sheet("confusion_matrix")
    if cm_img.exists():
        # taruh gambar di A1
        img = XLImage(str(cm_img))
        img.anchor = "A1"
        ws_cm.add_image(img)
        ws_cm["A20"] = "Catatan: gambar confusion_matrix.png diambil dari hasil evaluasi terakhir."
    else:
        add_df_to_sheet(ws_cm, pd.DataFrame([["confusion_matrix.png tidak ditemukan"]], columns=["info"]))

    # ===== Save to bytes =====
    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio.read()


# ===== UI =====
# Prasyarat minimal: metrics/report/cmatrix setidaknya 1 ada
if not (metrics_path.exists() or report_csv.exists() or cm_img.exists()):
    empty_state("Belum ada output evaluasi (metrics/report/confusion matrix). Jalankan Training & Evaluasi dulu.")
    st.stop()

gen = st.button("Generate laporan_sentimen.xlsx", type="primary")

if gen:
    try:
        with st.spinner("Membuat laporan Excel..."):
            data = generate_report_bytes()

            # simpan ke disk juga (opsional, untuk dashboard/export berikutnya)
            PROCESSED_DIR.mkdir(parents=True, exist_ok=True)
            out_xlsx.write_bytes(data)

        st.success("Laporan berhasil dibuat.")
        st.download_button(
            "Download laporan_sentimen.xlsx",
            data=data,
            file_name="laporan_sentimen.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    except PermissionError:
        st.error("Gagal menyimpan karena file sedang dibuka di Excel. Tutup Excel lalu coba Generate lagi.")
    except Exception as e:
        st.error(f"Gagal generate laporan: {e}")

# Kalau file sudah ada, tetap sediakan download (biar nggak hilang)
if out_xlsx.exists():
    st.markdown("---")
    st.caption("File laporan terakhir sudah tersedia.")
    st.download_button(
        "Download laporan_sentimen.xlsx (terakhir)",
        data=out_xlsx.read_bytes(),
        file_name="laporan_sentimen.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )